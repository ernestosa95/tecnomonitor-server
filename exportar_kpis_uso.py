#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Exporta KPIs de uso (RIS/PACS) a un Excel de tabla larga.

Salida: una fila por (hospital, mes, modalidad) con:
  - ordenes_ris_admitidos : suma de 'admitidos' del RIS
  - estudios_pacs_almacenados : suma de 'almacenados' del PACS

Las metricas NO estan en columnas SQL: viven dentro de
reportes_uso.kpi_json_data como JSON (application_metrics). Este script
parsea ese JSON replicando la logica de generator_report.py:
  - fecha del evento = start_time_extraction del JSON; si falta, timestamp de la fila
  - agrupacion mensual: YYYY-MM
  - exclusiones: AETs = CLIENT/WADO/PACS, Modalidades = DOC

Uso:
    python exportar_kpis_uso.py --db monitor_hospitales.db --salida kpis_uso.xlsx
    # opcional acotar fechas (inclusive/exclusivo sobre el mes o dia):
    python exportar_kpis_uso.py --desde 2025-01-01 --hasta 2025-07-01
"""
import argparse
import json
import sqlite3
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXCLUDED_AETS = {"CLIENT", "WADO", "PACS"}
EXCLUDED_MODS = {"DOC"}

# Hospitales a exportar. Dejar vacio = todos.
# Se puede fijar aca o pasar por --hospitales H001,H002
HOSPITALES = ["P01", "P02", "P03", "P04", "P05", "P06", "P07", "P08", "P09", "P10", "P11", "P12", "P13", "P14", "P15", "P16", "P17", "P18", "P19", "P20", "P21", "P22", "P23", "P24", "P25", "P26"]  # ej: ["H001", "H002"]


def parse_fecha_evento(metrics, ts_fila):
    """Fecha del evento: start_time_extraction si existe, si no el timestamp de la fila."""
    s = metrics.get("start_time_extraction")
    if s:
        try:
            return datetime.fromisoformat(s).replace(tzinfo=None)
        except (ValueError, TypeError):
            pass
    if ts_fila is None:
        return None
    try:
        if isinstance(ts_fila, str):
            return datetime.strptime(ts_fila[:19], "%Y-%m-%d %H:%M:%S")
        return ts_fila.replace(tzinfo=None)
    except (ValueError, TypeError):
        return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default="monitor_hospitales.db", help="Ruta a la base SQLite")
    ap.add_argument("--salida", default="kpis_uso.xlsx", help="Archivo Excel de salida")
    ap.add_argument("--desde", default=None, help="Fecha desde YYYY-MM-DD (inclusive)")
    ap.add_argument("--hasta", default=None, help="Fecha hasta YYYY-MM-DD (exclusivo)")
    ap.add_argument("--hospitales", default=None,
                    help="IDs de hospital separados por coma. Ej: H001,H002. "
                         "Vacio = todos (o usa la lista HOSPITALES del script).")
    args = ap.parse_args()

    # Filtro de hospitales: prioridad al parametro CLI, si no la lista del script
    if args.hospitales:
        hospitales_filtro = {h.strip() for h in args.hospitales.split(",") if h.strip()}
    else:
        hospitales_filtro = set(HOSPITALES)

    f_desde = datetime.strptime(args.desde, "%Y-%m-%d") if args.desde else None
    f_hasta = datetime.strptime(args.hasta, "%Y-%m-%d") if args.hasta else None

    con = sqlite3.connect(args.db)
    con.row_factory = sqlite3.Row

    # Nombre de hospital
    nombres = {}
    try:
        for r in con.execute("SELECT hospital_id, nombre FROM hospitales_metadata"):
            nombres[r["hospital_id"]] = r["nombre"]
    except sqlite3.OperationalError:
        pass  # si no existe la tabla, usamos el id

    # Acumuladores: clave (hospital_id, mes, modalidad)
    ris = defaultdict(int)   # admitidos
    pacs = defaultdict(int)  # almacenados
    claves = set()

    sql = "SELECT hospital_id, timestamp, kpi_json_data FROM reportes_uso"
    params = []
    if hospitales_filtro:
        marcadores = ",".join("?" * len(hospitales_filtro))
        sql += f" WHERE hospital_id IN ({marcadores})"
        params = list(hospitales_filtro)
    filas = con.execute(sql, params).fetchall()

    for row in filas:
        raw = row["kpi_json_data"]
        if not raw:
            continue
        try:
            metrics = json.loads(raw)
        except (json.JSONDecodeError, TypeError):
            continue

        fecha = parse_fecha_evento(metrics, row["timestamp"])
        if fecha is None:
            continue
        if f_desde and fecha < f_desde:
            continue
        if f_hasta and fecha >= f_hasta:
            continue

        mes = fecha.strftime("%Y-%m")
        hid = row["hospital_id"]

        # RIS -> admitidos
        for item in metrics.get("ris", []) or []:
            aet = item.get("aet")
            eq = item.get("equipo")
            mod = item.get("mod", "") or "S/D"
            nombre_ris = eq or aet or "Desc"
            if (nombre_ris in EXCLUDED_AETS or aet in EXCLUDED_AETS
                    or mod in EXCLUDED_MODS):
                continue
            k = (hid, mes, mod)
            ris[k] += int(item.get("admitidos", 0) or 0)
            claves.add(k)

        # PACS -> almacenados
        for item in metrics.get("pacs", []) or []:
            aet = item.get("aet")
            mod = item.get("mod", "") or "S/D"
            if aet in EXCLUDED_AETS or mod in EXCLUDED_MODS:
                continue
            k = (hid, mes, mod)
            pacs[k] += int(item.get("almacenados", 0) or 0)
            claves.add(k)

    con.close()

    # Aviso si algun ID pedido no trajo datos
    if hospitales_filtro:
        encontrados = {k[0] for k in claves}
        faltantes = hospitales_filtro - encontrados
        if faltantes:
            print("AVISO: sin datos para los IDs: " + ", ".join(sorted(faltantes)))

    # Armar filas ordenadas: hospital, mes, modalidad
    orden = sorted(claves, key=lambda k: (nombres.get(k[0], k[0]), k[1], k[2]))

    # Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPIs Uso"
    headers = ["hospital_id", "hospital", "mes", "modalidad",
               "ordenes_ris_admitidos", "estudios_pacs_almacenados"]
    ws.append(headers)

    for k in orden:
        hid, mes, mod = k
        ws.append([hid, nombres.get(hid, hid), mes, mod,
                   ris.get(k, 0), pacs.get(k, 0)])

    # Formato encabezado
    header_fill = PatternFill("solid", fgColor="0072B5")
    header_font = Font(bold=True, color="FFFFFF")
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    # Anchos y utilidades
    anchos = [14, 32, 10, 12, 22, 26]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(args.salida)
    print(f"OK -> {args.salida}  ({len(orden)} filas)")


if __name__ == "__main__":
    main()