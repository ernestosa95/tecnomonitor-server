#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Exportacion resumen por hospital (una fila por hospital).

Columnas:
  hospital_id, hospital, provincia,
  go_live               : primera fecha con ordenes RIS o estudios PACS
  estudios_almacenados  : suma de almacenados en PACS
  ordenes_admitidas     : suma RIS 'admitidos'
  ordenes_asociadas     : suma RIS 'con_imagen'  (asociadas)
  ordenes_definitivas   : suma RIS 'definitivos'
  estudios_ia           : almacenados de AETs que empiezan con 'ENT'
  equipos_almacenan     : cantidad de AE Titles distintos que almacenan
  tb_almacenados        : uso del disco J de la VM que termina en APPV (TB)
  tb_disponibles        : 60 TB (capacidad fija) - tb_almacenados
  uso_ram_pct           : promedio de uso de RAM del server (%)

Fuentes:
  - reportes_uso.kpi_json_data (RIS/PACS)  -> misma logica que generator_report
  - reportes_historicos.full_json_data     -> infraestructura (disco J, RAM)

No todos los hospitales tienen el 100% de los datos: lo que falte queda vacio
sin cortar la exportacion.

Uso:
    python exportar_resumen_hospitales.py --db monitor_hospitales.db --salida resumen.xlsx
    python exportar_resumen_hospitales.py --hospitales H001,H002 --desde 2025-01-01 --hasta 2025-08-01
"""
import argparse
import json
import sqlite3
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXCLUDED_AETS = {"CLIENT", "WADO", "PACS"}
EXCLUDED_MODS = {"DOC"}

PREFIJO_IA = "ENT"          # AE Titles de IA
CAPACIDAD_TB_TOTAL = 60.0   # capacidad fija por hospital
GB_POR_TB = 1024.0          # 1 TB = 1024 GB (binario)

HOSPITALES = []  # ej: ["H001", "H002"]; vacio = todos (o usar --hospitales)


def parse_fecha_evento(metrics, ts_fila):
    s = metrics.get("start_time_extraction")
    if s:
        try:
            return datetime.fromisoformat(s).replace(tzinfo=None)
        except (ValueError, TypeError):
            pass
    if ts_fila is None:
        return None
    try:
        if isinstance(ts_fila, str):
            return datetime.strptime(ts_fila[:19], "%Y-%m-%d %H:%M:%S")
        return ts_fila.replace(tzinfo=None)
    except (ValueError, TypeError):
        return None


def parse_ts(ts):
    try:
        if isinstance(ts, str):
            return datetime.strptime(ts[:19], "%Y-%m-%d %H:%M:%S")
        return ts.replace(tzinfo=None)
    except (ValueError, TypeError):
        return None


def uso_disco_j_appv_tb(full_json):
    """Uso (GB->TB) del disco J en la VM cuyo id termina en APPV. None si no esta."""
    try:
        data = json.loads(full_json) if isinstance(full_json, str) else full_json
    except (json.JSONDecodeError, TypeError):
        return None
    for vm in data.get("virtual_layer", []) or []:
        vid = str(vm.get("id", "")).upper()
        if not vid.endswith("APPV"):
            continue
        for disk in vm.get("storage", []) or []:
            mount = str(disk.get("mount_point", "")).strip().upper()
            if mount.startswith("J"):  # J, J:, J:\
                total = float(disk.get("total_gb") or 0)
                free = float(disk.get("free_gb") or 0)
                usado_gb = max(total - free, 0.0)
                return round(usado_gb / GB_POR_TB, 2)
    return None


def ram_pct(full_json):
    try:
        data = json.loads(full_json) if isinstance(full_json, str) else full_json
    except (json.JSONDecodeError, TypeError):
        return None
    tele = (data.get("physical_layer") or {}).get("telemetry") or {}
    val = (tele.get("ram") or {}).get("usage_percent")
    try:
        return float(val) if val is not None else None
    except (ValueError, TypeError):
        return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default="monitor_hospitales.db")
    ap.add_argument("--salida", default="resumen_hospitales.xlsx")
    ap.add_argument("--desde", default=None, help="YYYY-MM-DD (inclusive)")
    ap.add_argument("--hasta", default=None, help="YYYY-MM-DD (exclusivo)")
    ap.add_argument("--hospitales", default=None, help="IDs separados por coma")
    args = ap.parse_args()

    if args.hospitales:
        hfiltro = {h.strip() for h in args.hospitales.split(",") if h.strip()}
    else:
        hfiltro = set(HOSPITALES)

    f_desde = datetime.strptime(args.desde, "%Y-%m-%d") if args.desde else None
    f_hasta = datetime.strptime(args.hasta, "%Y-%m-%d") if args.hasta else None

    con = sqlite3.connect(args.db)
    con.row_factory = sqlite3.Row

    # Metadata
    nombres, provincias = {}, {}
    try:
        for r in con.execute(
                "SELECT hospital_id, nombre, provincia FROM hospitales_metadata"):
            nombres[r["hospital_id"]] = r["nombre"]
            provincias[r["hospital_id"]] = r["provincia"] or ""
    except sqlite3.OperationalError:
        pass

    def filtro_sql(col_ts):
        cond, params = [], []
        if hfiltro:
            cond.append("hospital_id IN (%s)" % ",".join("?" * len(hfiltro)))
            params += list(hfiltro)
        return (" WHERE " + " AND ".join(cond)) if cond else "", params

    # ---- reportes_uso: RIS / PACS ----
    acc = defaultdict(lambda: {
        "estudios": 0, "admitidos": 0, "asociadas": 0, "definitivas": 0,
        "ia": 0, "aets": set(), "go_live": None,
    })

    w, p = filtro_sql("timestamp")
    for row in con.execute(
            "SELECT hospital_id, timestamp, kpi_json_data FROM reportes_uso" + w, p):
        raw = row["kpi_json_data"]
        if not raw:
            continue
        try:
            metrics = json.loads(raw)
        except (json.JSONDecodeError, TypeError):
            continue
        fecha = parse_fecha_evento(metrics, row["timestamp"])
        if fecha is None:
            continue
        if f_desde and fecha < f_desde:
            continue
        if f_hasta and fecha >= f_hasta:
            continue

        hid = row["hospital_id"]
        a = acc[hid]
        hay_actividad = False

        for item in metrics.get("ris", []) or []:
            aet = item.get("aet")
            eq = item.get("equipo")
            mod = item.get("mod", "") or ""
            nombre_ris = eq or aet or "Desc"
            if (nombre_ris in EXCLUDED_AETS or aet in EXCLUDED_AETS
                    or mod in EXCLUDED_MODS):
                continue
            adm = int(item.get("admitidos", 0) or 0)
            aso = int(item.get("con_imagen", 0) or 0)
            de = int(item.get("definitivos", 0) or 0)
            a["admitidos"] += adm
            a["asociadas"] += aso
            a["definitivas"] += de
            if adm or aso or de or int(item.get("totales", 0) or 0):
                hay_actividad = True

        for item in metrics.get("pacs", []) or []:
            aet = (item.get("aet") or "Desc")
            mod = item.get("mod", "") or ""
            if aet in EXCLUDED_AETS or mod in EXCLUDED_MODS:
                continue
            val = int(item.get("almacenados", 0) or 0)
            if val <= 0:
                continue
            a["estudios"] += val
            a["aets"].add(aet)
            if aet.upper().startswith(PREFIJO_IA):
                a["ia"] += val
            hay_actividad = True

        if hay_actividad:
            if a["go_live"] is None or fecha < a["go_live"]:
                a["go_live"] = fecha

    # ---- reportes_historicos: infraestructura (disco J, RAM) ----
    # Streaming fila por fila: la tabla puede tener millones de snapshots,
    # asi que NO usamos fetchall() ni acumulamos filas en memoria.
    infra = defaultdict(lambda: {"tb": None, "tb_ts": None,
                                 "ram_sum": 0.0, "ram_n": 0})

    w, p = filtro_sql("timestamp")
    sql_infra = ("SELECT hospital_id, timestamp, full_json_data "
                 "FROM reportes_historicos" + w
                 + (" AND" if w else " WHERE") + " full_json_data IS NOT NULL")
    try:
        cur = con.execute(sql_infra, p)
    except sqlite3.OperationalError:
        cur = []

    for row in cur:  # iteracion perezosa: sqlite entrega de a lotes
        ts = parse_ts(row["timestamp"])
        if ts is None:
            continue
        if f_desde and ts < f_desde:
            continue
        if f_hasta and ts >= f_hasta:
            continue
        hid = row["hospital_id"]
        d = infra[hid]

        r = ram_pct(row["full_json_data"])
        if r is not None:
            d["ram_sum"] += r
            d["ram_n"] += 1

        tb = uso_disco_j_appv_tb(row["full_json_data"])
        if tb is not None and (d["tb_ts"] is None or ts >= d["tb_ts"]):
            d["tb"] = tb
            d["tb_ts"] = ts

    con.close()

    # ---- Armar filas ----
    hospitales = set(acc) | set(infra)
    if hfiltro:
        hospitales &= hfiltro

    registros = []
    for hid in hospitales:
        a = acc.get(hid, {})
        d = infra.get(hid, {})
        tb_alm = d.get("tb")
        tb_disp = round(CAPACIDAD_TB_TOTAL - tb_alm, 2) if tb_alm is not None else None
        ram_prom = round(d["ram_sum"] / d["ram_n"], 1) if d.get("ram_n") else None
        go = a.get("go_live")
        registros.append([
            hid,
            nombres.get(hid, hid),
            provincias.get(hid, ""),
            go.strftime("%Y-%m-%d") if go else "",
            a.get("estudios", 0),
            a.get("admitidos", 0),
            a.get("asociadas", 0),
            a.get("definitivas", 0),
            a.get("ia", 0),
            len(a.get("aets", set())),
            tb_alm,
            tb_disp,
            ram_prom,
        ])

    registros.sort(key=lambda r: r[1])

    if hfiltro:
        faltan = hfiltro - {r[0] for r in registros}
        if faltan:
            print("AVISO: sin datos para: " + ", ".join(sorted(faltan)))

    # ---- Excel ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen hospitales"
    headers = ["hospital_id", "hospital", "provincia", "go_live",
               "estudios_almacenados", "ordenes_admitidas", "ordenes_asociadas",
               "ordenes_definitivas", "estudios_ia", "equipos_almacenan",
               "tb_almacenados", "tb_disponibles", "uso_ram_pct"]
    ws.append(headers)
    for r in registros:
        ws.append(r)

    hf = PatternFill("solid", fgColor="0072B5")
    ff = Font(bold=True, color="FFFFFF")
    for c in ws[1]:
        c.fill = hf
        c.font = ff
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    anchos = [14, 30, 20, 12, 20, 18, 18, 18, 14, 18, 16, 16, 14]
    for i, wd in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(args.salida)
    print(f"OK -> {args.salida}  ({len(registros)} hospitales)")


if __name__ == "__main__":
    main()