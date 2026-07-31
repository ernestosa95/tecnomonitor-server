#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Inventario de equipos que ALMACENAN en el PACS, por hospital.

Se apoya en el almacenamiento PACS (no en el RIS), de modo que aparecen
tambien los equipos que almacenan pero no tienen flujo en el RIS.

Columnas de salida (una fila por hospital / AE Title / modalidad):
  - hospital_id
  - hospital
  - aet                 : AE Title del equipo
  - equipo              : nombre amigable (resuelto via RIS); vacio si no hay
  - modalidad
  - tipo                : 'IA' si el AET empieza con ENT_, si no 'Equipo medico'
  - tiene_flujo_ris     : 'Si' / 'No' (si el AET aparece en el RIS del hospital)
  - total_almacenados   : suma de estudios DICOM almacenados en el periodo
  - meses_activo        : cantidad de meses distintos con almacenamiento
  - primer_mes / ultimo_mes

Datos: reportes_uso.kpi_json_data (JSON, application_metrics), misma logica
de fecha y exclusiones que generator_report.py.

Uso:
    python exportar_equipos_pacs.py --db monitor_hospitales.db --salida equipos_pacs.xlsx
    python exportar_equipos_pacs.py --hospitales H001,H002 --desde 2025-01-01 --hasta 2025-07-01
"""
import argparse
import json
import sqlite3
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXCLUDED_AETS = {"CLIENT", "WADO", "PACS"}
EXCLUDED_MODS = {"DOC"}

# Tokens que NO son una modalidad de imagen real, sino objetos secundarios
# (documentos, structured reports, secondary capture, presentation state, key object).
# Se usan para derivar la modalidad principal de un equipo, no para excluir estudios.
MODS_SECUNDARIAS = {"DOC", "SR", "SC", "PR", "KO"}

# Hospitales a exportar. Vacio = todos (o usar --hospitales).
HOSPITALES = []  # ej: ["H001", "H002"]


def parse_fecha_evento(metrics, ts_fila):
    s = metrics.get("start_time_extraction")
    if s:
        try:
            return datetime.fromisoformat(s).replace(tzinfo=None)
        except (ValueError, TypeError):
            pass
    if ts_fila is None:
        return None
    try:
        if isinstance(ts_fila, str):
            return datetime.strptime(ts_fila[:19], "%Y-%m-%d %H:%M:%S")
        return ts_fila.replace(tzinfo=None)
    except (ValueError, TypeError):
        return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default="monitor_hospitales.db")
    ap.add_argument("--salida", default="equipos_pacs.xlsx")
    ap.add_argument("--desde", default=None, help="YYYY-MM-DD (inclusive)")
    ap.add_argument("--hasta", default=None, help="YYYY-MM-DD (exclusivo)")
    ap.add_argument("--hospitales", default=None,
                    help="IDs separados por coma. Vacio = todos.")
    args = ap.parse_args()

    if args.hospitales:
        hospitales_filtro = {h.strip() for h in args.hospitales.split(",") if h.strip()}
    else:
        hospitales_filtro = set(HOSPITALES)

    f_desde = datetime.strptime(args.desde, "%Y-%m-%d") if args.desde else None
    f_hasta = datetime.strptime(args.hasta, "%Y-%m-%d") if args.hasta else None

    con = sqlite3.connect(args.db)
    con.row_factory = sqlite3.Row

    nombres = {}
    provincias = {}
    try:
        for r in con.execute(
                "SELECT hospital_id, nombre, provincia FROM hospitales_metadata"):
            nombres[r["hospital_id"]] = r["nombre"]
            provincias[r["hospital_id"]] = r["provincia"] or ""
    except sqlite3.OperationalError:
        pass

    sql = "SELECT hospital_id, timestamp, kpi_json_data FROM reportes_uso"
    params = []
    if hospitales_filtro:
        marcadores = ",".join("?" * len(hospitales_filtro))
        sql += f" WHERE hospital_id IN ({marcadores})"
        params = list(hospitales_filtro)
    filas = con.execute(sql, params).fetchall()
    con.close()

    # aet -> equipo (resuelto desde RIS), por hospital
    dic_aet = defaultdict(dict)          # {hid: {aet: equipo}}
    ris_aets = defaultdict(set)          # {hid: {aets con flujo RIS}}
    # Acumulador PACS por (hid, aet) -> UN equipo por AE Title por hospital
    pacs = defaultdict(lambda: {"total": 0, "meses": set(),
                                "mod_w": defaultdict(int), "raw_top": ("", -1)})

    for row in filas:
        raw = row["kpi_json_data"]
        if not raw:
            continue
        try:
            metrics = json.loads(raw)
        except (json.JSONDecodeError, TypeError):
            continue

        fecha = parse_fecha_evento(metrics, row["timestamp"])
        if fecha is None:
            continue
        if f_desde and fecha < f_desde:
            continue
        if f_hasta and fecha >= f_hasta:
            continue

        mes = fecha.strftime("%Y-%m")
        hid = row["hospital_id"]

        # RIS: alimenta el diccionario de nombres y el set de AETs con flujo
        for item in metrics.get("ris", []) or []:
            aet = item.get("aet")
            eq = item.get("equipo")
            if aet and eq:
                dic_aet[hid][aet] = eq
            if aet:
                ris_aets[hid].add(aet)

        # PACS: base del inventario (una entrada por AE Title)
        for item in metrics.get("pacs", []) or []:
            aet = (item.get("aet") or "Desc")
            mod = item.get("mod", "") or "S/D"
            if aet in EXCLUDED_AETS or mod in EXCLUDED_MODS:
                continue
            val = int(item.get("almacenados", 0) or 0)
            if val <= 0:
                continue
            k = (hid, aet)
            rec = pacs[k]
            rec["total"] += val
            rec["meses"].add(mes)
            # La modalidad puede venir combinada, ej. "DX\\CR" o "MG\\SR\\DOC".
            # Repartimos el peso entre las modalidades de imagen reales.
            for tok in mod.split("\\"):
                tok = tok.strip().upper()
                if tok and tok not in MODS_SECUNDARIAS:
                    rec["mod_w"][tok] += val
            if val > rec["raw_top"][1]:
                rec["raw_top"] = (mod, val)

    # Armar filas: una por (hospital, aet)
    registros = []
    for (hid, aet), d in pacs.items():
        equipo = dic_aet[hid].get(aet, "")
        tipo = "IA" if aet.upper().startswith("ENT_") else "Equipo medico"
        flujo = "Si" if aet in ris_aets[hid] else "No"
        meses = sorted(d["meses"])
        # Modalidad principal = la de mayor almacenamiento; lista = todas.
        if d["mod_w"]:
            orden_mod = sorted(d["mod_w"].items(), key=lambda x: -x[1])
            modalidad = orden_mod[0][0]
            modalidades = ", ".join(m for m, _ in orden_mod)
        else:
            modalidad = d["raw_top"][0]  # fallback: solo tokens secundarios
            modalidades = modalidad
        registros.append([
            hid, nombres.get(hid, hid), provincias.get(hid, ""), aet, equipo,
            modalidad, modalidades, tipo, flujo, d["total"], len(meses),
            meses[0], meses[-1],
        ])

    # Orden: hospital, luego mayor almacenamiento primero
    registros.sort(key=lambda r: (r[1], -r[9]))

    if hospitales_filtro:
        faltantes = hospitales_filtro - {r[0] for r in registros}
        if faltantes:
            print("AVISO: sin datos PACS para: " + ", ".join(sorted(faltantes)))

    # Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipos PACS"
    headers = ["hospital_id", "hospital", "provincia", "aet", "equipo",
               "modalidad", "modalidades", "tipo", "tiene_flujo_ris",
               "total_almacenados", "meses_activo", "primer_mes", "ultimo_mes"]
    ws.append(headers)
    for r in registros:
        ws.append(r)

    header_fill = PatternFill("solid", fgColor="0072B5")
    header_font = Font(bold=True, color="FFFFFF")
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    # Resaltar los que almacenan pero NO tienen flujo RIS (columna I = 9)
    amarillo = PatternFill("solid", fgColor="FFF2CC")
    for i, r in enumerate(registros, start=2):
        if r[8] == "No":
            ws.cell(row=i, column=9).fill = amarillo

    anchos = [14, 30, 20, 18, 26, 12, 20, 16, 16, 20, 14, 12, 12]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(args.salida)
    print(f"OK -> {args.salida}  ({len(registros)} equipos)")


if __name__ == "__main__":
    main()
